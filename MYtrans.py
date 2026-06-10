import sys
import re
import sqlite3
from pathlib import Path
from datetime import datetime

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Side, Font, PatternFill


# ==========================================
# ① 設定ゾーン（ここだけ触ればOK）
# ==========================================
if getattr(sys, "frozen", False):
    BASE_DIR = Path(sys.executable).resolve().parent
else:
    BASE_DIR = Path(__file__).resolve().parent

CSV_FILE = BASE_DIR / "GRID_DATA.CSV"

BASE_OUTPUT_DIR = BASE_DIR / "MYtrans_output"
DB_FILE = BASE_OUTPUT_DIR / "mytrans_work.db"

OUTPUT_FILE_SUMMARY = "MYtrans_sum.xlsx"
OUTPUT_FILE_DETAIL = "MYtrans_dir.xlsx"
OUTPUT_FILE_PENDING = "未出荷リスト.xlsx"

TARGET_STATUS_PENDING = ["未出荷", "出荷中"]


# ==========================================
# ② 共通処理
# ==========================================
def clean_text(value) -> str:
    if pd.isna(value):
        return ""
    return str(value).replace("\ufeff", "").strip()


def normalize_status(value) -> str:
    s = clean_text(value).replace(" ", "").replace("　", "")

    mapping = {
        "出荷済": "出荷済み",
        "出荷済み": "出荷済み",
        "出荷中": "出荷中",
        "未出荷": "未出荷",
    }

    return mapping.get(s, s)


def read_csv_safely(file_path: Path) -> pd.DataFrame:
    print("読込対象:", file_path)

    with open(file_path, "rb") as f:
        header = f.read(4)

    # xlsx系ファイルは先頭が PK になる
    if header[:2] == b"PK":
        print("判定：Excel形式のファイルです。read_excelで読み込みます。")

        all_sheets = pd.read_excel(
            file_path,
            dtype=str,
            engine="openpyxl",
            sheet_name=None
        )

        print("シート一覧:", list(all_sheets.keys()))

        for sheet_name, df in all_sheets.items():
            print("-----")
            print("確認中シート:", sheet_name)
            print("行数:", len(df))
            print("列数:", len(df.columns))
            print("列名一覧:", df.columns.tolist())

            if len(df) > 0 and len(df.columns) > 0:
                print(f"使用シート決定: {sheet_name}")
                return df

        raise ValueError("Excel内にデータが入っているシートが見つかりません")

    encodings = [
        "cp932",
        "shift_jis",
        "utf-8-sig",
        "utf-8",
    ]

    seps = [
        ",",
        "\t",
        None,
    ]

    last_error = None

    for enc in encodings:
        for sep in seps:
            try:
                df = pd.read_csv(
                    file_path,
                    encoding=enc,
                    encoding_errors="replace",
                    dtype=str,
                    sep=sep,
                    engine="python",
                    on_bad_lines="warn"
                )

                print(f"CSV読込成功: encoding={enc} / sep={repr(sep)}")
                print("読込行数:", len(df))
                print("読込列数:", len(df.columns))
                print("列名一覧:", df.columns.tolist())

                return df

            except Exception as e:
                print(f"CSV読込失敗: encoding={enc} / sep={repr(sep)} / {e}")
                last_error = e

    raise ValueError(f"読み込み失敗: {file_path} / 最後のエラー: {last_error}")


def to_numeric_safe(series):
    return pd.to_numeric(
        series.astype(str)
        .str.replace(",", "", regex=False)
        .str.replace("kg", "", regex=False)
        .str.replace("KG", "", regex=False)
        .str.strip(),
        errors="coerce"
    ).fillna(0)


def normalize_columns(df: pd.DataFrame) -> pd.DataFrame:
    df.columns = [clean_text(c) for c in df.columns]

    df = df.rename(columns={
        "受注ＮＯ": "受注NO",
        "受注ＮＯ枝番": "受注NO枝番",
        "指示ＮＯ": "指示NO",

        "ルート": "ルート名",
        "配送ルート": "ルート名",
        "ルートコード": "ルート名",

        "納入先": "納入先名",
        "得意先名": "納入先名",
        "出荷先名": "納入先名",

        "数量": "受注員数",
        "員数": "受注員数",

        "重量": "受注重量",
    })

    return df


def clean_customer_name(value):
    s = clean_text(value)
    s = re.sub(r"\(.*?\)|（.*?）", "", s)
    return s.strip()


def load_dataframe_to_sqlite(df: pd.DataFrame, db_file: Path, table_name: str = "orders"):
    conn = sqlite3.connect(db_file)
    try:
        df.to_sql(table_name, conn, if_exists="replace", index=False)
        print(f"SQLite取込成功: {db_file} / table={table_name} / rows={len(df)}")
    finally:
        conn.close()


def read_orders_from_sqlite(db_file: Path) -> pd.DataFrame:
    conn = sqlite3.connect(db_file)
    try:
        return pd.read_sql_query("SELECT * FROM orders", conn)
    finally:
        conn.close()


# ==========================================
# ③ Excel共通スタイル
# ==========================================
def set_border(cell):
    thin = Side(style="thin", color="999999")
    cell.border = Border(top=thin, left=thin, right=thin, bottom=thin)


def apply_common_sheet_style(ws, orientation="landscape"):
    ws.sheet_view.showGridLines = False

    # A4
    ws.page_setup.paperSize = 9
    ws.page_setup.orientation = orientation

    # A4に収める
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0

    ws.sheet_properties.pageSetUpPr.fitToPage = True

    ws.page_margins.left = 0.3
    ws.page_margins.right = 0.3
    ws.page_margins.top = 0.5
    ws.page_margins.bottom = 0.5
    ws.page_margins.header = 0.2
    ws.page_margins.footer = 0.2


def safe_save_workbook(wb, output_file: Path):
    try:
        wb.save(output_file)
        print("保存成功:", output_file)
    except PermissionError:
        alt = output_file.parent / f"{output_file.stem}_{datetime.now().strftime('%H%M%S')}.xlsx"
        wb.save(alt)
        print("元ファイルが開いていたため別名保存:", alt)


# ==========================================
# ④ MYtrans_sum 出力
# 区分列なし
# ルート名 → 納入先名 の入れ子
# ==========================================
def export_mytrans_sum(df: pd.DataFrame, output_file: Path):
    wb = Workbook()
    ws = wb.active
    ws.title = "MYtrans_sum"

    # 4列なのでA4縦
    apply_common_sheet_style(ws, orientation="portrait")

    title_font = Font(size=16, bold=True)
    header_font = Font(size=11, bold=True, color="FFFFFF")
    route_font = Font(size=11, bold=True)
    normal_font = Font(size=10)
    total_font = Font(size=11, bold=True)

    header_fill = PatternFill("solid", fgColor="404040")
    route_fill = PatternFill("solid", fgColor="D9EAF7")
    total_fill = PatternFill("solid", fgColor="E2F0D9")

    center = Alignment(horizontal="center", vertical="center")
    left = Alignment(horizontal="left", vertical="center", shrink_to_fit=True)
    left_wrap = Alignment(horizontal="left", vertical="center", wrap_text=True)
    right = Alignment(horizontal="right", vertical="center")

    ws["A1"] = "MYtrans_sum"
    ws["A1"].font = title_font

    ws["A2"] = datetime.now().strftime("%Y年%m月%d日 %H:%M:%S")
    ws["A2"].font = normal_font

    headers = [
        "ルート名",
        "納入先名",
        "合計/受注員数",
        "合計/受注重量",
    ]

    start_row = 4

    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=start_row, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center
        set_border(cell)

    ws.column_dimensions["A"].width = 16
    ws.column_dimensions["B"].width = 42
    ws.column_dimensions["C"].width = 16
    ws.column_dimensions["D"].width = 18

    summary = (
        df.groupby(["ルート名", "納入先名"], dropna=False)
        .agg({
            "受注員数": "sum",
            "受注重量": "sum",
        })
        .reset_index()
        .sort_values(["ルート名", "納入先名"])
    )

    route_totals = (
        summary.groupby("ルート名", dropna=False)
        .agg({
            "受注員数": "sum",
            "受注重量": "sum",
        })
        .reset_index()
        .sort_values("ルート名")
    )

    row_no = start_row + 1

    for _, route_row in route_totals.iterrows():
        route_name = route_row["ルート名"]
        route_qty = route_row["受注員数"]
        route_weight = round(route_row["受注重量"])

        # ルート合計行
        values = [
            route_name,
            "ルート合計",
            route_qty,
            route_weight,
        ]

        for col_idx, value in enumerate(values, start=1):
            cell = ws.cell(row=row_no, column=col_idx, value=value)
            cell.font = route_font
            cell.fill = route_fill
            cell.alignment = center if col_idx in [1, 3, 4] else left
            set_border(cell)

        ws.row_dimensions[row_no].height = 22
        row_no += 1

        customer_rows = summary[summary["ルート名"] == route_name]

        for _, customer_row in customer_rows.iterrows():
            values = [
                "",
                customer_row["納入先名"],
                customer_row["受注員数"],
                round(customer_row["受注重量"]),
            ]

            for col_idx, value in enumerate(values, start=1):
                cell = ws.cell(row=row_no, column=col_idx, value=value)
                cell.font = normal_font
                cell.alignment = center if col_idx in [1, 3, 4] else left_wrap
                set_border(cell)

            ws.row_dimensions[row_no].height = 24
            row_no += 1

    grand_qty = summary["受注員数"].sum()
    grand_weight = round(summary["受注重量"].sum())

    # 総合計行：ルート名＋納入先名セルを結合
    ws.merge_cells(start_row=row_no, start_column=1, end_row=row_no, end_column=2)
    total_cell = ws.cell(row=row_no, column=1, value="総合計")
    total_cell.font = total_font
    total_cell.fill = total_fill
    total_cell.alignment = center
    set_border(total_cell)

    # 結合されたB列側にも罫線・塗りを入れる
    b_cell = ws.cell(row=row_no, column=2)
    b_cell.fill = total_fill
    set_border(b_cell)

    qty_cell = ws.cell(row=row_no, column=3, value=grand_qty)
    qty_cell.font = total_font
    qty_cell.fill = total_fill
    qty_cell.alignment = center
    set_border(qty_cell)

    weight_cell = ws.cell(row=row_no, column=4, value=grand_weight)
    weight_cell.font = total_font
    weight_cell.fill = total_fill
    weight_cell.alignment = center
    set_border(weight_cell)

    ws.row_dimensions[row_no].height = 24

    ws.freeze_panes = "A5"
    ws.auto_filter.ref = f"A4:D{row_no}"
    ws.print_area = f"A1:D{row_no}"

    safe_save_workbook(wb, output_file)


# ==========================================
# ⑤ MYtrans_dir 出力
# 区分列なし
# ルート名 → 納入先名 → 明細
# ==========================================
def export_mytrans_dir(df: pd.DataFrame, output_file: Path):
    wb = Workbook()
    ws = wb.active
    ws.title = "MYtrans_dir"

    # 商品名が長いためA4横
    apply_common_sheet_style(ws, orientation="landscape")

    title_font = Font(size=16, bold=True)
    header_font = Font(size=11, bold=True, color="FFFFFF")
    route_font = Font(size=11, bold=True)
    customer_font = Font(size=10, bold=True)
    normal_font = Font(size=10)

    header_fill = PatternFill("solid", fgColor="404040")
    route_fill = PatternFill("solid", fgColor="D9EAF7")
    customer_fill = PatternFill("solid", fgColor="FCE4D6")

    center = Alignment(horizontal="center", vertical="center")
    left = Alignment(horizontal="left", vertical="center", shrink_to_fit=True)
    left_wrap = Alignment(horizontal="left", vertical="center", wrap_text=True)

    ws["A1"] = "MYtrans_dir"
    ws["A1"].font = title_font

    ws["A2"] = datetime.now().strftime("%Y年%m月%d日 %H:%M:%S")
    ws["A2"].font = normal_font

    headers = [
        "ルート名",
        "納入先名",
        "受注NO",
        "商品名",
        "受注員数",
    ]

    start_row = 4

    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=start_row, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center
        set_border(cell)

    ws.column_dimensions["A"].width = 16
    ws.column_dimensions["B"].width = 32
    ws.column_dimensions["C"].width = 18
    ws.column_dimensions["D"].width = 54
    ws.column_dimensions["E"].width = 12

    df_detail = df[
        ["ルート名", "納入先名", "受注NO", "商品名", "受注員数"]
    ].copy()

    df_detail = df_detail.sort_values(
        ["ルート名", "納入先名", "受注NO", "商品名"]
    )

    route_totals = (
        df_detail.groupby("ルート名", dropna=False)["受注員数"]
        .sum()
        .reset_index()
        .sort_values("ルート名")
    )

    row_no = start_row + 1

    for _, route_row in route_totals.iterrows():
        route_name = route_row["ルート名"]
        route_qty = route_row["受注員数"]

        # ルート合計行
        values = [
            route_name,
            "ルート合計",
            "",
            "",
            route_qty,
        ]

        for col_idx, value in enumerate(values, start=1):
            cell = ws.cell(row=row_no, column=col_idx, value=value)
            cell.font = route_font
            cell.fill = route_fill
            cell.alignment = center if col_idx in [1, 5] else left
            set_border(cell)

        ws.row_dimensions[row_no].height = 22
        row_no += 1

        route_df = df_detail[df_detail["ルート名"] == route_name]

        customer_totals = (
            route_df.groupby("納入先名", dropna=False)["受注員数"]
            .sum()
            .reset_index()
            .sort_values("納入先名")
        )

        for _, customer_row in customer_totals.iterrows():
            customer_name = customer_row["納入先名"]
            customer_qty = customer_row["受注員数"]

            # 納入先合計行
            values = [
                "",
                customer_name,
                "",
                "納入先合計",
                customer_qty,
            ]

            for col_idx, value in enumerate(values, start=1):
                cell = ws.cell(row=row_no, column=col_idx, value=value)
                cell.font = customer_font
                cell.fill = customer_fill
                cell.alignment = center if col_idx in [1, 5] else left_wrap
                set_border(cell)

            ws.row_dimensions[row_no].height = 22
            row_no += 1

            detail_rows = route_df[route_df["納入先名"] == customer_name]

            for _, detail in detail_rows.iterrows():
                values = [
                    "",
                    "",
                    detail["受注NO"],
                    detail["商品名"],
                    detail["受注員数"],
                ]

                for col_idx, value in enumerate(values, start=1):
                    cell = ws.cell(row=row_no, column=col_idx, value=value)
                    cell.font = normal_font
                    cell.alignment = center if col_idx in [1, 5] else left_wrap
                    set_border(cell)

                ws.row_dimensions[row_no].height = 30
                row_no += 1

    ws.freeze_panes = "A5"
    ws.auto_filter.ref = f"A4:E{row_no}"
    ws.print_area = f"A1:E{row_no}"

    safe_save_workbook(wb, output_file)


# ==========================================
# ⑥ 未出荷リスト 出力
# 状態が 未出荷 / 出荷中 のもの
# ==========================================
def export_pending_list(df: pd.DataFrame, output_file: Path):
    wb = Workbook()
    ws = wb.active
    ws.title = "未出荷リスト"

    # 商品名が長いためA4横
    apply_common_sheet_style(ws, orientation="landscape")

    title_font = Font(size=16, bold=True)
    header_font = Font(size=11, bold=True, color="FFFFFF")
    normal_font = Font(size=10)

    header_fill = PatternFill("solid", fgColor="404040")

    center = Alignment(horizontal="center", vertical="center")
    left_wrap = Alignment(horizontal="left", vertical="center", wrap_text=True)

    ws["A1"] = "未出荷リスト"
    ws["A1"].font = title_font

    ws["A2"] = datetime.now().strftime("%Y年%m月%d日 %H:%M:%S")
    ws["A2"].font = normal_font

    headers = [
        "状態",
        "納入先名",
        "受注NO",
        "商品名",
        "受注員数",
    ]

    start_row = 4

    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=start_row, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center
        set_border(cell)

    ws.column_dimensions["A"].width = 12
    ws.column_dimensions["B"].width = 34
    ws.column_dimensions["C"].width = 18
    ws.column_dimensions["D"].width = 54
    ws.column_dimensions["E"].width = 12

    target_statuses = [normalize_status(x) for x in TARGET_STATUS_PENDING]

    df_pending = df[df["状態"].isin(target_statuses)].copy()

    df_pending = df_pending[
        ["状態", "納入先名", "受注NO", "商品名", "受注員数"]
    ].sort_values(["状態", "納入先名", "受注NO", "商品名"])

    row_no = start_row + 1

    if df_pending.empty:
        ws.cell(row=row_no, column=1, value="対象データなし")
        ws.cell(row=row_no, column=1).font = normal_font
    else:
        for _, row in df_pending.iterrows():
            values = [
                row["状態"],
                row["納入先名"],
                row["受注NO"],
                row["商品名"],
                row["受注員数"],
            ]

            for col_idx, value in enumerate(values, start=1):
                cell = ws.cell(row=row_no, column=col_idx, value=value)
                cell.font = normal_font
                cell.alignment = center if col_idx in [1, 5] else left_wrap
                set_border(cell)

            ws.row_dimensions[row_no].height = 30
            row_no += 1

    ws.freeze_panes = "A5"
    ws.auto_filter.ref = f"A4:E{max(row_no, start_row + 1)}"
    ws.print_area = f"A1:E{max(row_no, start_row + 1)}"

    safe_save_workbook(wb, output_file)


# ==========================================
# ⑦ メイン処理
# ==========================================
def main():
    BASE_OUTPUT_DIR.mkdir(parents=True, exist_ok=True)

    if not CSV_FILE.exists():
        raise FileNotFoundError(f"CSVファイルが見つかりません: {CSV_FILE}")

    df = read_csv_safely(CSV_FILE)
    df = normalize_columns(df)

    print("CSV列名一覧:", df.columns.tolist())

    required_cols = [
        "ルート名",
        "納入先名",
        "受注員数",
        "受注重量",
        "状態",
        "受注NO",
        "商品名",
    ]

    for col in required_cols:
        if col not in df.columns:
            raise ValueError(f"{col}列がありません")

    df["ルート名"] = df["ルート名"].apply(clean_text)
    df["納入先名"] = df["納入先名"].apply(clean_customer_name)
    df["状態"] = df["状態"].apply(normalize_status)
    df["受注NO"] = df["受注NO"].apply(clean_text)
    df["商品名"] = df["商品名"].apply(clean_text)

    df["受注員数"] = to_numeric_safe(df["受注員数"])
    df["受注重量"] = to_numeric_safe(df["受注重量"])

    df_sql = df[
        [
            "ルート名",
            "納入先名",
            "受注員数",
            "受注重量",
            "状態",
            "受注NO",
            "商品名",
        ]
    ].copy()

    load_dataframe_to_sqlite(df_sql, DB_FILE)

    df_orders = read_orders_from_sqlite(DB_FILE)

    export_mytrans_sum(
        df_orders,
        BASE_OUTPUT_DIR / OUTPUT_FILE_SUMMARY,
    )

    export_mytrans_dir(
        df_orders,
        BASE_OUTPUT_DIR / OUTPUT_FILE_DETAIL,
    )

    export_pending_list(
        df_orders,
        BASE_OUTPUT_DIR / OUTPUT_FILE_PENDING,
    )

    print("処理完了")


if __name__ == "__main__":
    main()